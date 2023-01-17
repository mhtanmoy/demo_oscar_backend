import openpyxl
from oscar.apps.catalogue.models import ProductClass, ProductImage
from apps.catalogue.models import Product, Category, ProductAttribute, ProductAttributeValue
from apps.partner.models import Zone
from apps.partner.models import Partner, StockRecord
from urllib.request import urlretrieve
from django.core.files import File
def parser(data):
    #print(data)
    product_class_id = ProductClass.objects.get(name = data["product_type"])

    category_id= Category.objects.get(name = data["catagories"])
    try:

        zone = data["zone"]
        zone_qs = Zone.objects.get(title=zone)
        if data["structure"] == "Child product":
            structure = "child"
            parent = Product.objects.get(id = data["parent"])

            created, _ = Product.objects.update_or_create(
                product_class = product_class_id,
                title = data["title"],
                # need to check with product category
            )
            created.structure = structure
            created.parent = parent
        else:
            created, _ = Product.objects.update_or_create(
                product_class = product_class_id,
                title = data["title"],
            )
        created.meta_title = data["meta_title"]
        created.description = data["description"]
        created.meta_description = data["meta_description"]
        created.is_discountable = data["is_discountable"]
        created.zone = zone_qs
        created.is_public= data["is_public"]
        created.save()
        created.categories.add(category_id)

        image, _ = urlretrieve(data["product_image"])

        #ProductImage().image_field.save(data["image_caption"], image )

        product_image_created, _ = ProductImage.objects.update_or_create(
            product = created,
            #original.save(data["image_caption"],image,save=True),
            caption = data["image_caption"] 
        )
        product_image_created.original = File(open(image,'rb'))
        #product_image_created.original = image
        product_image_created.save()

        attribute_name = data["attribute_name"].split(",")

        attribute_value = data["attribute_value"].split(",")

        
        for attribute_name_item in attribute_name:
            attribute_created, _ = ProductAttribute.objects.get_or_create(
                product = created,
                name = attribute_name_item,
            )

            for attribute_value_item in attribute_value:
                temp = attribute_value_item.split(":")
                if attribute_value_item == temp[0]:
                    attribute_value_created, _ = ProductAttributeValue.objects.get_or_create(
                        product = created,
                        attribute = attribute_created,
                        value_text = temp[1],
                    )

                    break
        
        partner = Partner.objects.filter(id = data["partner"]).first()

        stock_record_created, _ = StockRecord.objects.update_or_create(
            product = created,
            partner = partner,
            partner_sku = data["partner_sku"],
            price = data["price"],
            num_in_stock = data["num_in_stock"],
            num_allocated = data["num_allocated"],
            low_stock_threshold = data["low_stock_threshold"],
        )
        #print("Stock Record Created", stock_record_created)
        print(data["product_image"])
        print(data["image_caption"])


        #print("Product Image Created", product_image_created)


                    
        
    except Exception as e:
        print(e)
   

def run():

    # Give the location of the file
    path = r"D:\Amarshohor_1\amarshohor_backend\scripts\Product_Details.xlsx"

    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj["001"]
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row 
    # Loop will print all columns name
    for i in range(1, max_col + 1):
        cell_obj = sheet_obj.cell(row = 1, column = i)


    # Read data from Cells and store in dictionary

    for i in range(2, max_row+1):
        product_dict = {"structure":"","is_public":"","ups":"","parent":"",
        "title":"","slug":"","description":"","meta_title":"","meta_description":"",
        "product_type":"","attribute_name":"","attribute_value":"","product_options":"","recommended_products":"",
        "catagories":"","is_discountable":"","objects":"",
        "index":"","zone":"","unit":"","short_discription":"","partner":"","partner_sku":"","price":"","num_in_stock":"",
        "num_allocated":"","low_stock_threshold":"","product_image":"","image_caption":""}

        for j in range(1, max_col+1 ):
            cell_obj = sheet_obj.cell(row = i, column = j)
            product_dict[list(product_dict)[j-1]]  = cell_obj.value
        
        if product_dict["product_type"] is None:
            break

        #print(product_dict)

        parser(product_dict)
    # qs= ProductClass.objects.all()
    # print(qs)